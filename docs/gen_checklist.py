#!/usr/bin/env python3
"""
生成产品验收清单 Excel (亮色主题版)
基于 PRD 的所有功能、场景、边界条件
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws = wb.active
ws.title = "V1初验"

# ============================================================
# 颜色定义 - 亮色主题
# ============================================================
WHITE = 'FFFFFFFF'
LIGHT_GRAY = 'FFF5F5F5'
DARK_TEXT = 'FF222222'
MED_TEXT = 'FF555555'
NAVY = 'FF2F5496'
NAVY_LIGHT = 'FFD6E4F0'
BORDER = 'FFCCCCCC'

P0_BG = 'FFE8F5E9'
P1_BG = 'FFFFF8E1'
P2_BG = 'FFE3F2FD'

# ============================================================
# 样式定义
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=BORDER),
    right=Side(style='thin', color=BORDER),
    top=Side(style='thin', color=BORDER),
    bottom=Side(style='thin', color=BORDER),
)
align_center = Alignment(horizontal='center', vertical='center')
align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# 列定义
# ============================================================
headers = ['编号', '模块', '功能点', '优先级', '验收标准', '状态', '问题类型', '备注', 'DEV反馈修复状态', 'DEV备注']
col_widths = [10, 14, 34, 8, 55, 12, 14, 40, 20, 42]

# ============================================================
# 验收项目数据
# ============================================================
items = [
    # ── 导航与布局 ──
    ('M1', '导航与布局', '模块', '—', '—', '', '', '', '', ''),

    ('M1.1', '导航与布局', '侧栏 Tab 切换', 'P0',
     '4个Tab（今日热点/工具榜/人物动态/深度专题）点击切换，内容区正确响应', '', '', '', '', ''),
    ('M1.2', '导航与布局', '筛选器切换', 'P0',
     '最新情报/高热爆料/编辑精选 三个筛选按钮，点击后内容排序改变', '', '', '', '', ''),
    ('M1.3', '导航与布局', 'PC 端侧栏固定', 'P0',
     '≥768px 左侧栏 w-64 固定，右侧内容区自适应', '', '', '', '', ''),
    ('M1.4', '导航与布局', '移动端汉堡菜单', 'P0',
     '<768px 侧栏收为汉堡菜单，点击展开/关闭，遮罩层点击关闭', '', '', '', '', ''),
    ('M1.5', '导航与布局', '移动端横竖屏切换', 'P0',
     '横竖屏切换布局不崩溃，内容正常展示', '', '', '', '', ''),
    ('M1.6', '导航与布局', '爆料按钮', 'P0',
     '顶部栏右侧"爆料"按钮，点击跳转管理员登录页', '', '', '', '', ''),
    ('M1.7', '导航与布局', '页脚版权声明', 'P0',
     '底部展示版权信息和个人学习声明，邮箱 mikogao@qq.com', '', '', '', '', ''),
    ('M1.8', '导航与布局', '暗色模式', 'P0',
     '整体暗色主题：#0c0c0c 背景，#d4af37 金色强调，#ececeb 文字', '', '', '', '', ''),
    ('M1.9', '导航与布局', '字体加载', 'P0',
     '标题 Playfair Display 衬线斜体，正文 Inter，标签 JetBrains Mono', '', '', '', '', ''),

    # ── 内容卡片 ──
    ('M2', '内容卡片', '模块', '—', '—', '', '', '', '', ''),
    ('M2.1', '内容卡片', '卡片渲染（完整信息）', 'P0',
     '卡片展示：封面图 + 标题（衬线斜体）+ 摘要（2行截断）+ 标签 + 热度 + 来源 + 时间', '', '', '', '', ''),
    ('M2.2', '内容卡片', '卡片 Hover 效果', 'P0',
     'Hover 时边框变金色、卡片微上移（-translate-y-0.5）、过渡动画 300ms', '', '', '', '', ''),
    ('M2.3', '内容卡片', '新窗口跳转原文', 'P0',
     '点击卡片（除标签外）新窗口打开原文 URL', '', '', '', '', ''),
    ('M2.4', '内容卡片', '标签点击筛选', 'P0',
     '点击卡片标签 → 按该标签全局筛选内容', '', '', '', '', ''),
    ('M2.5', '内容卡片', '热度色标', 'P0',
     '热度值 ≥80 橙色、≥60 金色、≥40 灰色、<40 暗灰', '', '', '', '', ''),
    ('M2.6', '内容卡片', '热度显示格式', 'P0',
     '展示为 "● XX°C" 格式', '', '', '', '', ''),
    ('M2.7', '内容卡片', '图片暗色滤镜', 'P0',
     '有图：brightness-75 contrast-125 grayscale-[20%] 滤镜叠加', '', '', '', '', ''),
    ('M2.8', '内容卡片', '无图隐藏图片区', 'P0',
     '文章无配图时隐藏封面图区域（卡片变短）', '', '', '', '', ''),
    ('M2.9', '内容卡片', '数据源格式 "via xx"', 'P0',
     '来源显示 "via 数据源名称" 格式', '', '', '', '', ''),
    ('M2.10', '内容卡片', '管理员爆料标记', 'P0',
     '管理员录入文章卡片右上角红色"爆料"角标', '', '', '', '', ''),
    ('M2.11', '内容卡片', '骨架屏加载态', 'P0',
     '内容加载时展示 3 个骨架屏卡片（灰色脉冲动画）', '', '', '', '', ''),
    ('M2.12', '内容卡片', '空态提示', 'P0',
     '无内容时展示 "暂无内容，雷达扫描中…"', '', '', '', '', ''),
    ('M2.13', '内容卡片', '错误态与重试', 'P0',
     '加载失败时展示 "加载失败，点击重试" 按钮', '', '', '', '', ''),

    # ── 排版与筛选 ──
    ('M3', '排版与筛选', '模块', '—', '—', '', '', '', '', ''),
    ('M3.1', '排版与筛选', 'PC 端 3 列卡片布局', 'P0',
     '≥768px 三列瀑布流，卡片间距一致', '', '', '', '', ''),
    ('M3.2', '排版与筛选', '移动端单列布局', 'P0',
     '<768px 单列全宽卡片', '', '', '', '', ''),
    ('M3.3', '排版与筛选', '侧栏数据源筛选', 'P0',
     '侧栏展示可用数据源列表，点击筛选该源内容', '', '', '', '', ''),
    ('M3.4', '排版与筛选', '侧栏标签筛选', 'P0',
     '侧栏展示热门标签列表，点击筛选该标签内容', '', '', '', '', ''),
    ('M3.5', '排版与筛选', '中文内容筛选', 'P1',
     '提供"中文/全部"切换按钮，仅展示中文内容', '', '', '', '', ''),
    ('M3.6', '排版与筛选', '标签筛选激活态显示', 'P0',
     '按标签筛选时顶部显示当前标签和"清除"按钮', '', '', '', '', ''),
    ('M3.7', '排版与筛选', '编辑精选自动隐藏', 'P0',
     '编辑精选筛选器在数据为空时自动隐藏', '', '', '', '', ''),

    # ── 数据源 ──
    ('M4', '数据源', '模块', '—', '—', '', '', '', '', ''),
    ('M4.1', '数据源', 'GitHub Trending 抓取', 'P0',
     '通过 HTML 解析抓取 GitHub Trending 每日仓库，含名称/描述/星数/语言，≥10 条', '', '', '', '', ''),
    ('M4.2', '数据源', 'arXiv 论文抓取', 'P0',
     '通过 RSS 抓取 cs.AI 最新论文，含标题/摘要/作者/日期，≥10 条', '', '', '', '', ''),
    ('M4.3', '数据源', '机器之心资讯抓取', 'P0',
     '通过 RSS 抓取机器之心最新 AI 资讯', '', '', '', '', ''),
    ('M4.4', '数据源', '36氪资讯抓取', 'P1',
     '通过 RSS 抓取 36氪最新科技新闻', '', '', '', '', ''),
    ('M4.5', '数据源', "Lenny's Podcast 抓取", 'P0',
     '通过 Substack RSS 抓取播客内容，含封面/标题/摘要/音频链接', '', '', '', '', ''),
    ('M4.6', '数据源', '硅谷101 抓取', 'P0',
     '通过 Fireside RSS 抓取播客内容', '', '', '', '', ''),
    ('M4.7', '数据源', '播客 enclosure 解析降级', 'P0',
     '播客 enclosure 解析失败时退化为普通文章样式，不崩溃', '', '', '', '', ''),
    ('M4.8', '数据源', '备用 RSS 地址自动切换', 'P0',
     '主 RSS 地址失败时自动尝试 fallback_urls 中的备用地址', '', '', '', '', ''),
    ('M4.9', '数据源', '单源失败不影响其他', 'P0',
     '一个数据源抓取失败只记录日志，不影响其他源抓取和展示', '', '', '', '', ''),
    ('M4.10', '数据源', 'URL 去重', 'P0',
     '以 URL 为唯一约束，重复抓取自动跳过', '', '', '', '', ''),

    # ── 热门议题 ──
    ('M5', '热门议题', '模块', '—', '—', '', '', '', '', ''),
    ('M5.1', '热门议题', '深度专题 Tab 展示', 'P0',
     '「深度专题」Tab 展示热门议题卡片列表', '', '', '', '', ''),
    ('M5.2', '热门议题', '双行标题机制', 'P1',
     '议题卡片：主标题（热度最高文章标题）+ 副标题（"话题：#X · 共 Y 篇跨源探讨"）', '', '', '', '', ''),
    ('M5.3', '热门议题', '跨源聚合阈值', 'P1',
     '同一关键词在 ≥2 种源类型出现，且总文章 ≥3 篇', '', '', '', '', ''),
    ('M5.4', '热门议题', '48 小时聚合窗口', 'P1',
     '只聚合最近 48 小时抓取的内容', '', '', '', '', ''),
    ('M5.5', '热门议题', '置顶/头条机制', 'P1',
     '顶部大卡片：自动选 24h 内 hot_score 最高且有图文章，is_featured=true 可覆盖', '', '', '', '', ''),
    ('M5.6', '热门议题', '议题详情页', 'P1',
     '点击议题卡片跳转详情页，展示该话题下所有关联文章列表', '', '', '', '', ''),
    ('M5.7', '热门议题', '议题无数据降级', 'P1',
     '无聚合数据时隐藏顶部大卡片，仅展示常规议题列表或空态', '', '', '', '', ''),

    # ── 管理员 ──
    ('M6', '管理员', '模块', '—', '—', '', '', '', '', ''),
    ('M6.1', '管理员', 'Token 鉴权登录', 'P1',
     '访问 /admin 先显示 Token 输入框，输入正确 token 进入管理面板', '', '', '', '', ''),
    ('M6.2', '管理员', 'Token 失败 3 次锁定', 'P1',
     'Token 错误 3 次提示"请刷新页面重试"', '', '', '', '', ''),
    ('M6.3', '管理员', 'Session 级存储', 'P1',
     'Token 存入 sessionStorage，关闭页面即清除', '', '', '', '', ''),
    ('M6.4', '管理员', '爆料表单提交', 'P1',
     '表单字段：标题/原文链接/摘要/分类/配图/标签/是否置顶，提交写入 articles', '', '', '', '', ''),
    ('M6.5', '管理员', '爆料卡片标记', 'P1',
     '提交文章 is_admin_post=true，卡片带红色[爆料]角标', '', '', '', '', ''),
    ('M6.6', '管理员', '全量重新打标签', 'P1',
     '触发 POST /api/admin/retag，清空全部标签关联后重算', '', '', '', '', ''),
    ('M6.7', '管理员', '手动触发抓取', 'P1',
     '触发 POST /api/admin/fetch，执行一次全量 RSS 抓取', '', '', '', '', ''),
    ('M6.8', '管理员', '数据统计概览', 'P1',
     '管理面板展示：总文章数/24h新增/标签数/热门议题数/数据源数', '', '', '', '', ''),
    ('M6.9', '管理员', '标签概览', 'P1',
     '管理面板底部展示所有标签及其文章数', '', '', '', '', ''),

    # ── 后端功能 ──
    ('M7', '后端功能', '模块', '—', '—', '', '', '', '', ''),
    ('M7.1', '后端功能', 'RSS 定时抓取（4次/天）', 'P0',
     'node-cron 调度 UTC+8 8:00/12:00/18:00/22:00 执行全量抓取', '', '', '', '', ''),
    ('M7.2', '后端功能', '启动时自动抓取', 'P0',
     '服务启动后立即执行一次全量抓取（冷启动触发）', '', '', '', '', ''),
    ('M7.3', '后端功能', '顺序执行非并行', 'P0',
     'RSS 抓取单线程顺序遍历，不同时请求多个源', '', '', '', '', ''),
    ('M7.4', '后端功能', '热度评分计算', 'P0',
     '抓取时计算 hot_score = base_score × recency_boost（base: GitHub 65, arXiv 55, 机器之心 60, Lenny 70, 硅谷101 60）', '', '', '', '', ''),
    ('M7.5', '后端功能', '标签自动匹配', 'P0',
     '标题+摘要+content:encoded 前 200 字，\\bkeyword\\b 正则匹配，不区分大小写', '', '', '', '', ''),
    ('M7.6', '后端功能', '热门议题生成', 'P1',
     '每次抓取完成后自动触发 generateHotTopics() 重新聚合', '', '', '', '', ''),
    ('M7.7', '后端功能', 'API 分页查询', 'P0',
     '所有列表接口支持 page/limit 分页，默认 20 条，最大 50 条', '', '', '', '', ''),
    ('M7.8', '后端功能', '分页元数据', 'P0',
     'API 返回 pagination: { page, limit, total, totalPages }', '', '', '', '', ''),
    ('M7.9', '后端功能', '统一 JSON 响应格式', 'P0',
     '所有接口返回 { data, error, pagination? } 格式', '', '', '', '', ''),
    ('M7.10', '后端功能', '参数化查询防注入', 'P0',
     '所有 SQL 使用 $1, $2 参数化查询，禁止拼接 SQL', '', '', '', '', ''),
    ('M7.11', '后端功能', 'CORS 跨域配置', 'P0',
     '后端配置 cors() 允许前端域名跨域请求', '', '', '', '', ''),
    ('M7.12', '后端功能', '日志记录', 'P0',
     '每次抓取记录：时间/源名称/结果/新增条数/耗时', '', '', '', '', ''),

    # ── 非功能性需求 ──
    ('M8', '非功能需求', '模块', '—', '—', '', '', '', '', ''),
    ('M8.1', '非功能需求', '首屏加载 < 3s', 'P0',
     'Vercel CDN + 静态资源缓存，首屏加载时间 < 3s', '', '', '', '', ''),
    ('M8.2', '非功能需求', 'API 响应 < 200ms', 'P0',
     'API 查询（有索引）响应时间 < 200ms', '', '', '', '', ''),
    ('M8.3', '非功能需求', '响应式 320~1920px', 'P0',
     '320px ~ 1920px+ 自适应，无横向滚动', '', '', '', '', ''),
    ('M8.4', '非功能需求', '版权合规', 'P0',
     '只展示标题与摘要（<200字），新窗口跳转原文，页脚版权声明', '', '', '', '', ''),
    ('M8.5', '非功能需求', '浏览器兼容', 'P0',
     'Chrome/Firefox/Safari/Edge 最新版 + iOS Safari/Android Chrome', '', '', '', '', ''),
    ('M8.6', '非功能需求', 'User-Agent 声明', 'P0',
     '所有外部 HTTP 请求设置 User-Agent: Singularity-Radar/1.0', '', '', '', '', ''),
    ('M8.7', '非功能需求', 'UTC 存储/UTC+8 显示', 'P0',
     '数据库时间按 UTC 存储，前端展示转换为 UTC+8（北京时区）', '', '', '', '', ''),
    ('M8.8', '非功能需求', '内容缓存 72 小时', 'P0',
     'CACHE_TTL=72 小时，内容在缓存有效期内不重复抓取', '', '', '', '', ''),
]

today = datetime.date.today()
today_str = today.strftime('%Y-%m-%d')
today_num = today.strftime('%Y%m%d')

# ============================================================
# 写入 Excel
# ============================================================

# ── Title row 1 ──
ws.merge_cells('A1:J1')
cell = ws.cell(row=1, column=1)
cell.value = 'Singularity Radar · 产品验收清单'
cell.font = Font(name='Inter', size=16, bold=True, color=DARK_TEXT)
cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
cell.alignment = align_center
for c in range(2, 11):
    ws.cell(row=1, column=c).fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
ws.row_dimensions[1].height = 36

# ── Version row 2 ──
ws.merge_cells('A2:E2')
cell = ws.cell(row=2, column=1, value=f'版本 {today_num} ｜ 生成日期 {today_str}')
cell.font = Font(name='Inter', size=10, color=MED_TEXT)
cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
cell.alignment = Alignment(horizontal='left', vertical='center')
for c in range(2, 11):
    ws.cell(row=2, column=c).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
ws.row_dimensions[2].height = 24

# ── Status legend row 3 ──
ws.merge_cells('A3:J3')
cell = ws.cell(row=3, column=1,
               value='验收状态：通过 ｜ 不通过 ｜ 部分通过 ｜ 待重测 ｜ 跳过')
cell.font = Font(name='Inter', size=10, color=MED_TEXT)
cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
cell.alignment = align_center
for c in range(2, 11):
    ws.cell(row=3, column=c).fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
ws.row_dimensions[3].height = 22

# ── Type + DEV legend row 4 ──
ws.merge_cells('A4:J4')
cell = ws.cell(row=4, column=1,
               value='问题类型：BUG ｜ 体验优化 ｜ 需求变更　　　　DEV反馈：已修复 ｜ 非BUG待确认 ｜ 待修复 ｜ 需确认')
cell.font = Font(name='Inter', size=10, color=MED_TEXT)
cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
cell.alignment = align_center
for c in range(2, 11):
    ws.cell(row=4, column=c).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
ws.row_dimensions[4].height = 22

# ── Time row 5 ──
ws.merge_cells('A5:J5')
cell = ws.cell(row=5, column=1, value=f'验收时间：________ （手动录入）')
cell.font = Font(name='Inter', size=10, color=MED_TEXT)
cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
cell.alignment = align_center
for c in range(2, 11):
    ws.cell(row=5, column=c).fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
ws.row_dimensions[5].height = 22

# ── Header row 6 ──
header_font = Font(name='Inter', size=10, bold=True, color='FFFFFFFF')
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border
ws.row_dimensions[6].height = 30

# ── Column widths ──
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Data rows ──
module_fill = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type='solid')
module_font = Font(name='Inter', size=10, bold=True, color='FF1F3864')
cell_font = Font(name='Inter', size=10, color=DARK_TEXT)

p0_fill = PatternFill(start_color=P0_BG, end_color=P0_BG, fill_type='solid')
p1_fill = PatternFill(start_color=P1_BG, end_color=P1_BG, fill_type='solid')
p2_fill = PatternFill(start_color=P2_BG, end_color=P2_BG, fill_type='solid')

row = 7
for item in items:
    is_header = (item[2] == '模块')

    if is_header:
        for c in range(1, 11):
            cell = ws.cell(row=row, column=c)
            cell.value = item[0] if c == 1 else None
            cell.font = module_font if c == 1 else Font(name='Inter', size=10, color=DARK_TEXT)
            cell.fill = module_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 26
        row += 1
        continue

    priority = item[3]
    if priority == 'P0':
        row_fill = p0_fill
    elif priority == 'P1':
        row_fill = p1_fill
    else:
        row_fill = p2_fill

    for c in range(1, 11):
        cell = ws.cell(row=row, column=c, value=item[c - 1])
        cell.font = cell_font
        cell.fill = row_fill
        cell.border = thin_border
        if c in (1, 4, 6, 7):
            cell.alignment = align_center
        elif c in (3, 5, 8, 9, 10):
            cell.alignment = align_left_wrap
        else:
            cell.alignment = Alignment(vertical='center')

    ws.row_dimensions[row].height = 32
    row += 1

# ── Data validation dropdowns ──
dv_status = DataValidation(type='list', formula1='"通过,不通过,部分通过,待重测,跳过"', allow_blank=True)
dv_status.add(f'F7:F{row - 1}')
ws.add_data_validation(dv_status)

dv_type = DataValidation(type='list', formula1='"BUG,体验优化,需求变更"', allow_blank=True)
dv_type.add(f'G7:G{row - 1}')
ws.add_data_validation(dv_type)

dv_dev = DataValidation(type='list', formula1='"已修复,非BUG待确认,待修复,需确认"', allow_blank=True)
dv_dev.add(f'I7:I{row - 1}')
ws.add_data_validation(dv_dev)

# ── Excel Table (for robust filtering) ──
ref = f'A6:J{row - 1}'
table = Table(displayName='验收表', ref=ref)
table.tableStyleInfo = TableStyleInfo(
    name='TableStyleMedium2',
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=False, showColumnStripes=False,
)
ws.add_table(table)

# ── Freeze (below header row) ──
ws.freeze_panes = 'A7'

# ── Print ──
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# ── Save ──
output_path = f'/Users/miko/Desktop/AI文档/6.热搜/Singularity-Radar/docs/验收清单_{today_num}.xlsx'
wb.save(output_path)
print(f'✅ 已生成：{output_path}')
print(f'   共 {row - 7} 条验收项（含模块标题）')
print(f'   亮色主题 · 10 列 · 下拉菜单 · Excel 表格')
